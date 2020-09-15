from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

load_wb = load_workbook("./222.xlsx", data_only=True)

load_ws = load_wb['Sheet1']

cells = load_ws['B1':'CC12']

# 기준 기업
# for row in range(2, 3):
#     for col in range(2, 10):
#         print(load_ws.cell(column=col, row=row).value)


# 거래처
for row in range(2, 3):
    for col in range(10, 30):
        print(load_ws.cell(column=col, row=row).value)


load_wb.close()